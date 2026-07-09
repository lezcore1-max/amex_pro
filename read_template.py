"""
Read and describe the submission template structure exactly.
"""
import openpyxl
import os

TEMPLATE_PATH = r"c:\Users\vansh\Music\amex\campus_challenge_r1_submission_template.xlsx"

wb = openpyxl.load_workbook(TEMPLATE_PATH)
print(f"Sheet names: {wb.sheetnames}")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{'='*60}")
    print(f"Sheet: '{sheet_name}'")
    print(f"  Max row: {ws.max_row}, Max col: {ws.max_column}")
    # Print first 40 rows
    print("  First 40 rows:")
    for i, row in enumerate(ws.iter_rows(max_row=40, values_only=True)):
        print(f"    Row {i+1}: {row}")
    if ws.max_row > 40:
        print(f"  ... ({ws.max_row - 40} more rows) ...")
        # Print last 3 rows
        last_rows = list(ws.iter_rows(min_row=ws.max_row-2, values_only=True))
        for i, row in enumerate(last_rows):
            print(f"    Row {ws.max_row-2+i}: {row}")

wb.close()

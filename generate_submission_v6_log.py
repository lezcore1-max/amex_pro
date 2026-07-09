import pandas as pd
import numpy as np
import openpyxl

print("Loading data for V6 Log-Transformed P&L...")
df = pd.read_csv('campus_challenge_r1_data.csv')

features = [f'f{i}' for i in range(1, 24)]
df_imp = df.copy()

# Median imputation for missing values
for f in features:
    if df_imp[f].isnull().any():
        df_imp[f] = df_imp[f].fillna(df_imp[f].median())

print("Calculating Log-Transformed Revenue and Cost...")

# 1. GRANULAR REVENUE (Using specific category margins)
rev_raw = (df_imp['f6'] * 0.022 + df_imp['f9'] * 0.020 + df_imp['f10'] * 0.020 + 
           df_imp['f8'] * 0.015 + df_imp['f7'] * 0.010 + df_imp['f1'] * 0.18)

# 2. GRANULAR COST
cost_raw = (df_imp['f21'] * 0.01 + df_imp['f4'] * 0.005 + 
            df_imp['f11'] * df_imp['f17'] * 0.50 + 
            df_imp['f13'] * 30 + df_imp['f14'] + df_imp['f15'] * 10 + df_imp['f16'] + 
            df_imp['f2'] * 50 + df_imp['f3'] * 150)

# 3. LOG TRANSFORMATION
# np.log1p safely computes log(1 + x) for positive numbers. We use np.maximum to prevent negative values from creating NaNs.
log_rev = np.log1p(np.maximum(0, rev_raw))
log_cost = np.log1p(np.maximum(0, cost_raw))

# 4. LOG MARGIN SCORE
# Log(R) - Log(C) is mathematically equivalent to Log(R / C)
# This scores users by Profit Margin Ratio instead of Absolute Raw Dollars
margin_score = log_rev - log_cost
margin_score = margin_score.fillna(margin_score.median())

# Rank the final score
df['Prediction'] = margin_score.rank(pct=True).fillna(0.5)

# Write to official template
template_path = 'campus_challenge_r1_submission_template.xlsx'
print(f"Loading template: {template_path}")
wb = openpyxl.load_workbook(template_path)

ws_pred = wb['Predictions']
print("Writing 500,000 predictions (this will take 1-2 minutes)...")
for i, row in enumerate(df.itertuples(), start=2):
    ws_pred.cell(row=i, column=1, value=int(row.id))
    ws_pred.cell(row=i, column=2, value=float(row.Prediction))

ws_fw = wb['Profitability Framework']
framework_text = {
    'Variables Used': 'f1, f2, f3, f4, f6, f7, f8, f9, f10, f11, f13, f14, f15, f16, f17, f21',
    'Profitability Equation': 'Prediction = Rank( Log1p(Revenue) - Log1p(Cost) )',
    'Prediction Logic': 'We compute exact raw dollars for Revenue and Cost using standard banking coefficients (18% APR, interchange rates, LGD). Instead of simple subtraction (which gets dominated by whale outliers) or percentile ranking (which destroys dollar ratios), we apply a Log1p transformation. Mathematically, Log(R) - Log(C) is equivalent to Log(R / C), meaning this model strictly scores customers based on their Profit Margin Ratio.',
    'Variable Selection Logic': 'Used core financial features: category spends (f6-f10) and revolve (f1) for revenue; rewards (f4, f21), benefits (f13-f16), and credit loss (f11*f17) for costs.',
    'Coefficient/Weight Derivation': 'Category-specific interchange rates (1.0% to 2.2%), standard revolve APR (18%), rewards liability (1.0 cent/point), and Expected Credit Loss (Risk_Score * Line * 50% LGD).',
    'Feature Transformations': 'Missing values were median-imputed. Revenue and Cost were aggressively compressed using np.log1p() to handle the power-law skew inherent in credit card portfolios. The final margin score was percentile-ranked.',
    'Business Logic': 'In corporate analytics, absolute profit is heavily skewed by a few massive clients. Measuring "Profitability" as a margin (Log Revenue / Cost) identifies highly efficient cardmembers regardless of their absolute limit or spend level. A customer costing $100 to generate $200 (Margin 2.0) is fundamentally more "profitable" relative to risk than a whale costing $1M to generate $1.1M (Margin 1.1).',
    'Assumptions': 'Assumes the hidden target label normalizes for scale, valuing efficiency (margin) over pure volume. Assumes standard banking constants apply.',
    'Validation Approach': 'Tested against V3 (Raw Dollars) and V1 (Ranked Dollars). Log transformation mathematically bridges the gap, preserving relative dollar ratios while taming extreme outliers.',
    'Additional Notes (Optional)': 'Log transformation is the industry gold standard for handling heavily skewed financial metrics before continuous evaluation (like RMSE or AUC/Spearman correlation).'
}

for row_idx in range(2, 12):
    section_name = ws_fw.cell(row=row_idx, column=1).value
    if section_name in framework_text:
        ws_fw.cell(row=row_idx, column=2, value=framework_text[section_name])

out_path = 'amex_submission_round1_v6.xlsx'
wb.save(out_path)
print(f"Saved to {out_path}")

print("\n--- Verification ---")
wb_check = openpyxl.load_workbook(out_path, read_only=True)
ws_check = wb_check['Predictions']
rows = list(ws_check.iter_rows(values_only=True))

print(f"Sheet names match template: {wb_check.sheetnames == ['Predictions', 'Profitability Framework']}")
print(f"Total rows (incl header): {len(rows)}")
print(f"Header: {rows[0]}")
last_row = rows[-1]
print(f"Last row: {last_row}")

preds = [r[1] for r in rows[1:]]
ids = [r[0] for r in rows[1:]]
preds_clean = [float(p) for p in preds if p is not None and str(p).lower() != 'nan']
print(f"Any NaN/Inf: {len(preds_clean) != len(preds) or np.isnan(preds_clean).any() or np.isinf(preds_clean).any()}")
print(f"IDs unique: {len(set(ids)) == 500000}")
print(f"Framework filled: {ws_check.parent['Profitability Framework'].cell(row=2, column=2).value is not None}")

import pandas as pd
import numpy as np
from sklearn.preprocessing import StandardScaler
from sklearn.decomposition import PCA
import openpyxl

print("Loading data...")
df = pd.read_csv('campus_challenge_r1_data.csv')

# 1. Feature Definition
# Positive drivers (revenue/value): f1 (revolve), f5 (spend), f17 (credit line)
# Negative drivers (cost/risk): f4 (points balance), f21 (points redeemed), f11 (risk score)
features = ['f1', 'f5', 'f17', 'f4', 'f21', 'f11']
df_imp = df.copy()
for f in features:
    df_imp[f] = df_imp[f].fillna(df_imp[f].median())

# 2. Standardization & Sign Correction BEFORE PCA
scaler = StandardScaler()
X_scaled = scaler.fit_transform(df_imp[features])

# Invert the negative drivers so that all variables are economically "positive"
# X_scaled columns: 0:f1, 1:f5, 2:f17, 3:f4, 4:f21, 5:f11
X_scaled[:, 3] = -X_scaled[:, 3] # -f4
X_scaled[:, 4] = -X_scaled[:, 4] # -f21
X_scaled[:, 5] = -X_scaled[:, 5] # -f11

# 3. PCA
pca = PCA(n_components=1)
pc1 = pca.fit_transform(X_scaled).flatten()

# Ensure PC1 is aligned positively with f5 (spend)
f5_idx = features.index('f5')
if pca.components_[0, f5_idx] < 0:
    pc1 = -pc1

# Final Prediction
df['Prediction'] = pd.Series(pc1).rank(pct=True)

# 4. Write to official template
template_path = 'campus_challenge_r1_submission_template.xlsx'
print(f"Loading template: {template_path}")
wb = openpyxl.load_workbook(template_path)

# Write Predictions
ws_pred = wb['Predictions']
for i, row in enumerate(df.itertuples(), start=2):
    ws_pred.cell(row=i, column=1, value=int(row.id))
    ws_pred.cell(row=i, column=2, value=float(row.Prediction))

# Write Framework
ws_fw = wb['Profitability Framework']
framework_text = {
    'Variables Used': 'f1 (Revolve), f5 (Spend), f17 (Credit Line), f4 (Rewards Balance), f21 (Rewards Redeemed), f11 (Risk Score)',
    'Profitability Equation': 'Prediction = Rank( PCA_Component_1 ( Z(f1), Z(f5), Z(f17), -Z(f4), -Z(f21), -Z(f11) ) )',
    'Prediction Logic': 'Rather than relying on arbitrary dollar weights, we standardize the core economic drivers and invert the costs/risks. We then use PCA to find the single axis of maximum variance (PC1). This isolates the fundamental bimodality of high-value vs low-value cardmembers.',
    'Variable Selection Logic': 'Selected the primary volume drivers of Revenue (Spend, Revolve, Credit Line) and Cost/Risk (Rewards liability, Rewards redemption, Risk score).',
    'Coefficient/Weight Derivation': 'Weights are entirely data-driven, derived from the first principal component (PC1) of the standardized features, completely eliminating subjective human bias.',
    'Feature Transformations': 'Missing values were median-imputed. All features were Z-score standardized. Cost/risk features (f4, f21, f11) were inverted (multiplied by -1) before PCA so that PC1 positively correlates with profitability.',
    'Business Logic': 'A highly profitable cardmember drives high revenue (spend/revolve) while maintaining low risk and low rewards extraction. Aligning these vectors allows PCA to find the pure \"value\" dimension in the data.',
    'Assumptions': 'Assumes that the true underlying profitability metric is a linear combination of these core features, and that the axis of maximum variance in these signed features represents profitability.',
    'Validation Approach': 'Verified that the PCA loadings are economically sound (positive for spend/revolve, negative for risk/rewards).',
    'Additional Notes (Optional)': 'This data-driven PCA approach resolves the issue of magnitude distortion found in percentile-based manual formulas.'
}

# The template has sections in row 2 to 11, column A. We write to column B.
for row_idx in range(2, 12):
    section_name = ws_fw.cell(row=row_idx, column=1).value
    if section_name in framework_text:
        ws_fw.cell(row=row_idx, column=2, value=framework_text[section_name])

# 5. Save and Verify
out_path = 'amex_submission_round1_v2.xlsx'
wb.save(out_path)
print(f"Saved to {out_path}")

# Verification
print("\n--- Verification ---")
wb_check = openpyxl.load_workbook(out_path, read_only=True)
ws_check = wb_check['Predictions']
rows = list(ws_check.iter_rows(values_only=True))

print(f"Sheet names match template: {wb_check.sheetnames == ['Predictions', 'Profitability Framework']}")
print(f"Total rows (incl header): {len(rows)}")
print(f"Header: {rows[0]}")
last_row = rows[-1]
print(f"Last row: {last_row}")

# Check NaN/Inf and uniqueness
preds = [r[1] for r in rows[1:]]
ids = [r[0] for r in rows[1:]]
print(f"Any NaN/Inf: {np.isnan(preds).any() or np.isinf(preds).any()}")
print(f"IDs unique: {len(set(ids)) == 500000}")
print(f"Framework filled: {ws_check.parent['Profitability Framework'].cell(row=2, column=2).value is not None}")

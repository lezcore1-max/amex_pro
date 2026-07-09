import pandas as pd
import numpy as np
import openpyxl

print("Loading data...")
df = pd.read_csv('campus_challenge_r1_data.csv')

# Missing value imputation - median is safest for numericals
features = [f'f{i}' for i in range(1, 24)]
df_imp = df.copy()
for f in features:
    if df_imp[f].isnull().any():
        df_imp[f] = df_imp[f].fillna(df_imp[f].median())

print("Calculating Raw P&L...")

# 1. REVENUE
# Transaction Revenue: 2% interchange on Total Spend (f5)
# Interest Revenue: 18% APR on Revolve Balance (f1)
df_imp['revenue_raw'] = (df_imp['f5'] * 0.02) + (df_imp['f1'] * 0.18)

# 2. COST
# Rewards: 1 cent per point redeemed (f21) and 0.5 cents per point in balance (f4, representing future liability)
rewards_cost = (df_imp['f21'] * 0.01) + (df_imp['f4'] * 0.005)

# Credit Loss: Risk Score (f11) * Total Lend Line (f17) * 50% Loss Given Default
ecl_cost = df_imp['f11'] * df_imp['f17'] * 0.50

# Benefits Cost: Lounge accesses (f13 * $30), Cab usage (f15 * $10), Air credit (f14), Ent credit (f16)
benefits_cost = (df_imp['f13'] * 30.0) + df_imp['f14'] + (df_imp['f15'] * 10.0) + df_imp['f16']

# Servicing/Attrition Cost: Cancellation calls (f2 * $50), Collection calls (f3 * $150)
servicing_cost = (df_imp['f2'] * 50.0) + (df_imp['f3'] * 150.0)

df_imp['cost_raw'] = rewards_cost + ecl_cost + benefits_cost + servicing_cost

# 3. RAW NET PROFIT
df_imp['profit_raw'] = df_imp['revenue_raw'] - df_imp['cost_raw']

# 4. FINAL PREDICTION (Rank the raw dollars)
df['Prediction'] = df_imp['profit_raw'].rank(pct=True)

# 5. Write to official template
template_path = 'campus_challenge_r1_submission_template.xlsx'
print(f"Loading template: {template_path}")
wb = openpyxl.load_workbook(template_path)

# Write Predictions
ws_pred = wb['Predictions']
print("Writing 500,000 predictions (this may take a minute)...")
for i, row in enumerate(df.itertuples(), start=2):
    ws_pred.cell(row=i, column=1, value=int(row.id))
    ws_pred.cell(row=i, column=2, value=float(row.Prediction))

# Write Framework
ws_fw = wb['Profitability Framework']
framework_text = {
    'Variables Used': 'f1, f2, f3, f4, f5, f11, f13, f14, f15, f16, f17, f21',
    'Profitability Equation': 'Prediction = Rank( (f5*0.02 + f1*0.18) - (f21*0.01 + f4*0.005 + f11*f17*0.5 + f13*30 + f14 + f15*10 + f16 + f2*50 + f3*150) )',
    'Prediction Logic': 'We compute a strict Net Profit & Loss (P&L) in actual dollars for every cardmember. Revenues (interchange, interest) are summed, and costs (rewards, credit loss, benefits, servicing) are subtracted. Crucially, this arithmetic is performed in raw dollars *before* ranking, ensuring magnitude is preserved.',
    'Variable Selection Logic': 'Used only direct economic drivers. Ignored category spends (f6-f10) as they sum to more than f5, indicating double-counting or supplementary inclusions. f5 provides a cleaner overall spend base.',
    'Coefficient/Weight Derivation': 'Constants represent standard real-world financial proxies: 2% interchange, 18% APR, 1 cent/point redemption, 50% LGD on credit line exposure, and flat unit costs for lounge ($30) and calls ($50-$150).',
    'Feature Transformations': 'Missing values median-imputed. Final dollar profit was percentile-ranked (0 to 1) as required by the continuous scoring rule.',
    'Business Logic': 'The most profitable cardmembers drive massive absolute revenue through revolve interest and spend, while strictly minimizing their rewards extraction and carrying minimal default risk.',
    'Assumptions': 'Assumes f5 (Total Spend) is the true base for interchange. Assumes f1 (Revolve Balance) accrues interest at an average 18% APR.',
    'Validation Approach': 'Confirmed that the top 20% quintile is heavily dominated by revolve balance (the primary economic driver) while maintaining low risk profiles.',
    'Additional Notes (Optional)': 'This raw P&L formulation solves the magnitude distortion error present in models that rank revenue and cost components individually before subtracting.'
}

for row_idx in range(2, 12):
    section_name = ws_fw.cell(row=row_idx, column=1).value
    if section_name in framework_text:
        ws_fw.cell(row=row_idx, column=2, value=framework_text[section_name])

out_path = 'amex_submission_round1_v3.xlsx'
wb.save(out_path)
print(f"Saved to {out_path}")

# 6. Verification
print("\n--- Verification ---")
wb_check = openpyxl.load_workbook(out_path, read_only=True)
ws_check = wb_check['Predictions']
rows = list(ws_check.iter_rows(values_only=True))

print(f"Sheet names match template: {wb_check.sheetnames == ['Predictions', 'Profitability Framework']}")
print(f"Total rows (incl header): {len(rows)}")
print(f"Header: {rows[0]}")
last_row = rows[-1]
print(f"Last row: {last_row}")

preds = [r[1] for r in rows[1:]]
ids = [r[0] for r in rows[1:]]
print(f"Any NaN/Inf: {np.isnan(preds).any() or np.isinf(preds).any()}")
print(f"IDs unique: {len(set(ids)) == 500000}")
print(f"Framework filled: {ws_check.parent['Profitability Framework'].cell(row=2, column=2).value is not None}")
